from odoo import models, api
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

_logger = logging.getLogger(__name__)

class ReportSavingStateXLSX(models.AbstractModel):
    _name = 'report.planes_ahorro.report_saving_state_xlsx'
    _inherit = 'report.report_xlsx.abstract'  # Heredar de la clase correcta
    _description = 'Reporte de Estado de Cuenta en Excel'

    def create_xlsx_report(self, docids, data):
        """
        Método principal para generar el reporte en Excel.
        """
        docs = self.env['account.saving'].browse(docids)
        company = self.env.company

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"

        # Estilos
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        right_alignment = Alignment(horizontal='right')

        # Encabezados
        ws['A1'] = "Estado de Cuenta del Plan de Ahorro"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:G1')

        # Información del socio
        ws['A3'] = "Identificación:"
        ws['B3'] = docs.partner_id.vat or ""
        ws['A4'] = "Socio:"
        ws['B4'] = docs.partner_id.name

        # Información del plan
        ws['A6'] = "Código del plan:"
        ws['B6'] = docs.name
        ws['A7'] = "Tipo de ahorro:"
        ws['B7'] = "Normal" if docs.saving_type == 'normal' else "Balón"
        ws['A8'] = "Importe del Plan:"
        ws['B8'] = docs.saving_amount
        ws['A9'] = "Moneda:"
        ws['B9'] = docs.currency_id.name

        # Fechas y estado
        ws['A11'] = "Fecha de inicio:"
        ws['B11'] = docs.start_date
        ws['A12'] = "Fecha de Vencimiento:"
        ws['B12'] = docs.end_date
        ws['A13'] = "Estado:"
        ws['B13'] = docs.state_plan_description
        ws['A14'] = "Inscripción:"
        ws['B14'] = docs.serv_inscription_amount

        # Líneas del plan de ahorro
        ws['A16'] = "LÍNEAS DEL PLAN DE AHORRO"
        ws['A16'].font = header_font
        ws['A16'].alignment = center_alignment
        ws.merge_cells('A16:G16')

        # Encabezados de la tabla
        headers = ["#", "Fecha Vencimiento", "Fecha Pago", "Importe", "Valor Pagado", "Pendiente", "Estado"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=17, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment

        # Llenar la tabla con las líneas del plan
        row = 18
        payment_amount_total = 0
        pending_amount_total = 0
        for ml in docs.quota_ids.sorted(key=lambda l: l.number):
            ws.cell(row=row, column=1, value=ml.number)
            ws.cell(row=row, column=2, value=ml.date or "")
            ws.cell(row=row, column=3, value=ml.last_payment_date or "")
            ws.cell(row=row, column=4, value=ml.serv_admin_amount if ml.sequence == 0 else ml.saving_amount)
            ws.cell(row=row, column=5, value=ml.pagos)
            ws.cell(row=row, column=6, value=ml.pendiente)
            ws.cell(row=row, column=7, value=ml.estado_pago)
            payment_amount_total += ml.pagos
            pending_amount_total += ml.pendiente
            row += 1

        # Totales
        ws.cell(row=row, column=4, value="TOTAL").font = header_font
        ws.cell(row=row, column=5, value=payment_amount_total).font = header_font
        ws.cell(row=row, column=6, value=pending_amount_total).font = header_font

        # Guardar el archivo en un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read(), 'xlsx'