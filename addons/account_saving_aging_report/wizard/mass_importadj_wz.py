from odoo import _, api, fields, models
import base64
import io
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)

try:
    import xlsxwriter
except ImportError:
    _logger.warning("xlsxwriter not available")
    xlsxwriter = None

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl not available")
    openpyxl = None


class MassImportAdjWz(models.TransientModel):
    _name = 'mass.importadj.wz'
    _description = 'Mass Import Adjudication Wizard'

    file_data = fields.Binary(
        string='Archivo Excel',
        required=True,
        help='Subir el archivo Excel con los datos'
    )
    file_name = fields.Char(
        string='Nombre del Archivo',
        default='adjudicacion_import.xlsx'
    )
    state = fields.Selection([
        ('download', 'Descargar Plantilla'),
        ('import', 'Importar Datos'),
    ], default='download', string='Estado')
    
    # Campos para mostrar resultados
    total_records = fields.Integer(string='Total Registros', default=0)
    updated_records = fields.Integer(string='Registros Actualizados', default=0)
    error_records = fields.Integer(string='Registros con Error', default=0)
    result_message = fields.Text(string='Mensaje de Resultado', readonly=True)

    def action_download_excel(self):
        """Descarga el Excel con los datos de account.saving que cumplen las condiciones"""
        try:
            # Obtener los datos
            savings = self.env['account.saving'].search([
                ('state_plan', 'in', ['adjudicated', 'adjudicated_with_assets', 'adjudicated_without_assets'])
            ])
            
            if not savings:
                raise ValueError('No se encontraron registros con estado adjudicado')
            
            # Crear archivo Excel en memoria
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            
            # Formatos
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4CAF50',
                'color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            date_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'num_format': 'dd/mm/yyyy'
            })
            
            # Hoja principal
            worksheet = workbook.add_worksheet('Adjudicaciones')
            
            # Encabezados
            headers = ['ID', 'Nombre', 'Cliente', 'Fecha de Adjudicación']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
            
            # Escribir datos
            row = 1
            for saving in savings:
                worksheet.write(row, 0, saving.id, cell_format)
                worksheet.write(row, 1, saving.name or '', cell_format)
                worksheet.write(row, 2, saving.partner_id.name or '', cell_format)
                
                # Fecha - con formato
                if saving.adjudicate_date:
                    worksheet.write(row, 3, saving.adjudicate_date, date_format)
                else:
                    worksheet.write(row, 3, '', cell_format)
                
                row += 1
            
            # Ajustar ancho de columnas
            worksheet.set_column('A:A', 10)
            worksheet.set_column('B:B', 30)
            worksheet.set_column('C:C', 40)
            worksheet.set_column('D:D', 20)
            
            # Hoja de instrucciones
            instruction_ws = workbook.add_worksheet('Instrucciones')
            instruction_ws.write(0, 0, 'INSTRUCCIONES PARA LA IMPORTACIÓN', header_format)
            instruction_ws.write(1, 0, '1. Complete la columna "Fecha de Adjudicación" con el formato dd/mm/yyyy')
            instruction_ws.write(2, 0, '2. No modifique la columna ID')
            instruction_ws.write(3, 0, '3. Guarde el archivo y súbalo en el wizard')
            instruction_ws.set_column('A:A', 60)
            
            workbook.close()
            
            # Preparar para descarga
            output.seek(0)
            file_data = base64.b64encode(output.read())
            
            # Crear attachment
            attachment = self.env['ir.attachment'].create({
                'name': f'adjudicaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                'type': 'binary',
                'datas': file_data,
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            
            # Actualizar estado
            self.write({
                'state': 'import',
                'total_records': len(savings),
                'result_message': f'Archivo generado con {len(savings)} registros. Complete las fechas y suba el archivo para importar.'
            })
            
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }
            
        except Exception as e:
            _logger.error(f"Error generando Excel: {str(e)}")
            raise ValueError(f"Error al generar el Excel: {str(e)}")

    def action_import_excel(self):
        """Importa el archivo Excel y actualiza las fechas de adjudicación"""
        if not self.file_data:
            raise ValueError('Debe subir un archivo Excel')
        
        try:
            # Decodificar el archivo
            file_content = base64.b64decode(self.file_data)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            # Verificar encabezados
            headers = []
            for col in range(1, 5):  # Columnas A-D
                cell_value = worksheet.cell(row=1, column=col).value
                headers.append(cell_value)
            
            expected_headers = ['ID', 'Nombre', 'Cliente', 'Fecha de Adjudicación']
            for i, expected in enumerate(expected_headers):
                if headers[i] != expected:
                    raise ValueError(f'Encabezado incorrecto en columna {i+1}. Se esperaba "{expected}" y se encontró "{headers[i]}"')
            
            # Procesar datos
            updated = 0
            errors = 0
            error_messages = []
            
            for row in range(2, worksheet.max_row + 1):
                try:
                    # Leer datos
                    saving_id = worksheet.cell(row=row, column=1).value
                    adjudicate_date = worksheet.cell(row=row, column=4).value
                    
                    # Validar ID
                    if not saving_id:
                        error_messages.append(f'Fila {row}: ID vacío')
                        errors += 1
                        continue
                    
                    # Validar fecha
                    if not adjudicate_date:
                        error_messages.append(f'Fila {row} (ID {saving_id}): Fecha de adjudicación vacía')
                        errors += 1
                        continue
                    
                    # Convertir fecha si es string
                    if isinstance(adjudicate_date, str):
                        try:
                            # Intentar varios formatos
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']:
                                try:
                                    adjudicate_date = datetime.strptime(adjudicate_date, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            if isinstance(adjudicate_date, str):
                                raise ValueError('Formato de fecha no reconocido')
                        except Exception:
                            error_messages.append(f'Fila {row} (ID {saving_id}): Formato de fecha inválido. Use dd/mm/yyyy')
                            errors += 1
                            continue
                    
                    # Buscar y actualizar el registro
                    saving = self.env['account.saving'].search([
                        ('id', '=', int(saving_id)),
                        ('state_plan', 'in', ['adjudicated', 'adjudicated_with_assets', 'adjudicated_without_assets'])
                    ])
                    
                    if not saving:
                        error_messages.append(f'Fila {row} (ID {saving_id}): Registro no encontrado o no está en estado adjudicado')
                        errors += 1
                        continue
                    
                    # Actualizar fecha
                    saving.write({
                        'adjudicate_date': adjudicate_date
                    })
                    updated += 1
                    
                except Exception as e:
                    error_messages.append(f'Fila {row}: {str(e)}')
                    errors += 1
                    _logger.error(f"Error en fila {row}: {str(e)}")
            
            # Mostrar resultados
            result_msg = f"""
            ✅ PROCESO COMPLETADO
            
            📊 Total de registros procesados: {worksheet.max_row - 1}
            ✅ Registros actualizados: {updated}
            ❌ Registros con error: {errors}
            
            {chr(10).join(error_messages[:10])}
            {f'... y {len(error_messages) - 10} errores más' if len(error_messages) > 10 else ''}
            """
            
            self.write({
                'updated_records': updated,
                'error_records': errors,
                'result_message': result_msg,
                'state': 'download'
            })
            
            # Mensaje de éxito
            if errors == 0:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': '✅ Importación Exitosa',
                        'message': f'Se actualizaron {updated} registros correctamente',
                        'type': 'success',
                        'sticky': False,
                    }
                }
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': '⚠️ Importación Parcial',
                        'message': f'Se actualizaron {updated} registros, {errors} errores. Revise el mensaje de resultado.',
                        'type': 'warning',
                        'sticky': True,
                    }
                }
            
        except Exception as e:
            _logger.error(f"Error importando Excel: {str(e)}")
            raise ValueError(f"Error al importar el Excel: {str(e)}")

    def action_reset(self):
        """Resetear el wizard para una nueva operación"""
        self.write({
            'file_data': False,
            'file_name': 'adjudicacion_import.xlsx',
            'state': 'download',
            'total_records': 0,
            'updated_records': 0,
            'error_records': 0,
            'result_message': False
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }