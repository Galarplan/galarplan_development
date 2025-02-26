# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

import base64
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

from odoo import _, api, fields, models
from odoo.exceptions import ValidationError

class L10nEcMassWithholdImport(models.TransientModel):
    _name = 'l10n_ec.mass.withhold.import'
    _description = 'Mass Withhold Import'

    data_file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def action_import(self):
        self.ensure_one()
        if not self.data_file:
            raise ValidationError(_("Please upload a file to import."))

        # Decodificar el archivo
        file_content = base64.b64decode(self.data_file)
        file_stream = BytesIO(file_content)

        # Cargar el archivo XLSX
        try:
            workbook = load_workbook(filename=file_stream)
            sheet = workbook.active
        except Exception as e:
            raise ValidationError(_("The file is not a valid Excel file. Error: %s") % str(e))

        # Leer las filas del archivo
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]  # La primera fila contiene los encabezados
        required_headers = ['invoice_number', 'partner_id', 'date', 'tax_ids', 'base',]
        if not all(header in headers for header in required_headers):
            raise ValidationError(_("The Excel file must contain the following headers: %s") % ', '.join(required_headers))

        # Procesar cada fila
        for row in rows[1:]:  # Ignorar la fila de encabezados
            row_data = dict(zip(headers, row))
            self._process_row(row_data)

        return {'type': 'ir.actions.act_window_close'}

    def _process_row(self, row):
        # Obtener los valores de la fila
        invoice_number = row.get('invoice_number')
        partner_id = row.get('partner_id')
        date = row.get('date')
        tax_ids = row.get('tax_ids', '').split(',')  # Separar por comas
        base = row.get('base', '')  # Separar por comas

        
        # Buscar la factura
        invoice = self.env['account.move'].search([('l10n_latam_document_number', '=', invoice_number)], limit=1)
        if not invoice:
            raise ValidationError(_("Invoice %s not found.") % invoice_number)

        # Buscar el partner
        partner = invoice.partner_id.id
        if not partner:
            raise ValidationError(_("Partner with ID %s not found.") % partner_id)

        # Preparar las líneas de retención
        withhold_lines = []
        for tax_id in tax_ids:
            tax = self.env['account.tax'].search([('name','=',tax_id)],limit=1)
            
            if not tax:
                raise ValidationError(_("Tax with ID %s not found.") % tax_id.strip())
            base = base.replace('.', '').replace(',', '.')
            mount = float(base) * (tax.amount/100) * -1
            print('=====================',base,tax.amount,date)
            withhold_lines.append((0, 0, {
                'invoice_id': invoice.id,
                'tax_id': tax.id,
                'base': float(base),
                'amount': mount ,
            }))

        # Crear la retención
        withhold_vals = {
            'partner_id': partner,
            'date': date,
            'related_invoice_ids': [(6, 0, [invoice.id])],
            'withhold_line_ids': withhold_lines,
        }
        self.env['l10n_ec.wizard.account.withhold'].create(withhold_vals).action_create_and_post_withhold()